from openpyxl import load_workbook


def guardar_datos_excel(volumen1, volumen2, volumen3, volumen4,
                        probabilidades_rutas):
    # Cargar el archivo Excel
    workbook = load_workbook('entradas_simulador-normal.xlsx')
    sheet = workbook.active
    
    # Obtener la última fila ocupada
    last_row = sheet.max_row + 1
    
    # Guardar los datos en las columnas correspondientes
    sheet.cell(row=last_row, column=1).value = volumen1
    sheet.cell(row=last_row, column=2).value = volumen2
    sheet.cell(row=last_row, column=3).value = volumen3
    sheet.cell(row=last_row, column=4).value = volumen4


    for i, probabilidad in enumerate(probabilidades_rutas):
        sheet.cell(row=last_row, column=5 + i).value = probabilidad

    """ Descomentar para el primer escenario
    for i, probabilidad in enumerate(probabilidades_rutas2):
        sheet.cell(row=last_row, column=7 + i).value = probabilidad
    for i, probabilidad in enumerate(probabilidades_rutas3):
        sheet.cell(row=last_row, column=9 + i).value = probabilidad
    for i, probabilidad in enumerate(probabilidades_rutas4):
        sheet.cell(row=last_row, column=11 + i).value = probabilidad
    for i, probabilidad in enumerate(probabilidades_rutas5):
        sheet.cell(row=last_row, column=13 + i).value = probabilidad
    for i, probabilidad in enumerate(probabilidades_rutas6):
        sheet.cell(row=last_row, column=15 + i).value = probabilidad
    """

    # Guardar el archivo Excel
    workbook.save('entradas_simulador-normal.xlsx')


def guardar_datos_excel_salida(salidas):
    # Cargar el archivo Excel
    workbook = load_workbook('salidas_simulador-normal.xlsx')
    sheet = workbook.active
    
    # Obtener la última fila ocupada
    last_row = sheet.max_row + 1
    
    
    for i, probabilidad in enumerate(salidas):
        sheet.cell(row=last_row, column=1 + i).value = probabilidad
    
    # Guardar el archivo Excel
    workbook.save('salidas_simulador-normal.xlsx')